import io
import math
from dataclasses import asdict
from datetime import UTC, datetime
from typing import Annotated, Any

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import delete, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.scout_utils import NIGHTS_BADGE_COUNTS, is_nights_badge_count
from app.core.siie_import import map_row
from app.db import get_session
from app.deps import AdminUser, CurrentUser
from app.models import Profile, Scout, ScoutNightsBadge
from app.models.enums import OrdemSection
from app.schemas.scout import NightsBadgeOut, ScoutOut

router = APIRouter(prefix="/scouts", tags=["scouts"])

# camelCase request key -> snake_case model attribute (optional string fields).
_OPTIONAL_FIELDS = {
    "numeroAssociado": "numero_associado",
    "sexo": "sexo",
    "cc": "cc",
    "nif": "nif",
    "email": "email",
    "telefone": "telefone",
    "telemovel": "telemovel",
    "morada": "morada",
    "localidade": "localidade",
    "codigoPostal": "codigo_postal",
    "paiNome": "pai_nome",
    "paiTelefone": "pai_telefone",
    "paiEmail": "pai_email",
    "maeNome": "mae_nome",
    "maeTelefone": "mae_telefone",
    "maeEmail": "mae_email",
    "encarregadoNome": "encarregado_nome",
    "encarregadoTelefone": "encarregado_telefone",
    "encarregadoEmail": "encarregado_email",
}

_VALID_SECTIONS = {s.value for s in OrdemSection}


def _optional_string(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    return value.strip() or None


def _parse_date(value: Any) -> datetime | None:
    if not isinstance(value, str) or not value:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.astimezone(UTC).replace(tzinfo=None) if parsed.tzinfo else parsed


@router.get("")
async def list_scouts(
    user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    section: Annotated[str | None, Query()] = None,
    includeInactive: Annotated[bool, Query()] = False,  # noqa: N803
) -> dict[str, list[ScoutOut]]:
    conditions = []
    if section and section in _VALID_SECTIONS:
        conditions.append(Scout.section == OrdemSection(section))
    if not includeInactive:
        conditions.append(Scout.active.is_(True))
    rows = (
        await session.scalars(
            select(Scout)
            .where(*conditions)
            .order_by(Scout.section.asc(), Scout.last_name.asc(), Scout.first_name.asc())
        )
    ).all()
    return {"scouts": [ScoutOut.model_validate(s) for s in rows]}


@router.post("")
async def create_scout(
    admin: AdminUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    body: Annotated[dict[str, Any], Body()],
) -> dict[str, ScoutOut]:
    first_name = body["firstName"].strip() if isinstance(body.get("firstName"), str) else ""
    last_name = body["lastName"].strip() if isinstance(body.get("lastName"), str) else ""
    if not first_name or not last_name:
        raise HTTPException(status_code=400, detail="Nome e apelido obrigatórios")

    date_of_birth = _parse_date(body.get("dateOfBirth"))
    if not date_of_birth:
        raise HTTPException(status_code=400, detail="Data de nascimento inválida")

    section: OrdemSection | None = None
    raw_section = body.get("section")
    if raw_section is not None and raw_section != "":
        if raw_section not in _VALID_SECTIONS:
            raise HTTPException(status_code=400, detail="Secção inválida")
        section = OrdemSection(raw_section)

    joined_at = _parse_date(body.get("joinedAt")) or datetime.now(UTC).replace(tzinfo=None)
    raw_noites: Any = body.get("noitesCampoInicial")
    try:
        noites = float(raw_noites)
        noites_inicial = max(0, math.floor(noites)) if math.isfinite(noites) else 0
    except (TypeError, ValueError):
        noites_inicial = 0

    scout = Scout(
        first_name=first_name,
        last_name=last_name,
        date_of_birth=date_of_birth,
        section=section,
        joined_at=joined_at,
        noites_campo_inicial=noites_inicial,
    )
    for camel, attr in _OPTIONAL_FIELDS.items():
        setattr(scout, attr, _optional_string(body.get(camel)))

    session.add(scout)
    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=409, detail="Nº de associado já existe") from None
    await session.refresh(scout)
    return {"scout": ScoutOut.model_validate(scout)}


@router.post("/import")
async def import_scouts(
    admin: AdminUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    file: Annotated[UploadFile | None, File()] = None,
) -> dict[str, Any]:
    if file is None:
        raise HTTPException(status_code=400, detail="Ficheiro não fornecido")

    try:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        if not workbook.worksheets:
            raise HTTPException(status_code=400, detail="Folha vazia")
        sheet = workbook.worksheets[0]
        it = sheet.iter_rows(values_only=True)
        header = list(next(it, []))
        records = [
            dict(zip(header, row, strict=False)) for row in it if any(v is not None for v in row)
        ]
    except HTTPException:
        raise
    except Exception as err:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Não foi possível ler o ficheiro: {err}"
        ) from err

    summary: dict[str, Any] = {
        "total": len(records),
        "created": 0,
        "updated": 0,
        "linkedToProfile": 0,
        "errors": [],
    }

    valid = []
    for i, record in enumerate(records):
        mapped = map_row(record, i + 2)
        if not mapped.ok:
            summary["errors"].append(
                {"row": mapped.row, "nome": mapped.nome, "error": mapped.error}
            )
        else:
            valid.append(mapped)

    if not valid:
        return {"summary": summary}

    nins = [m.nin for m in valid]
    existing_nins = set(
        (
            await session.scalars(
                select(Scout.numero_associado).where(Scout.numero_associado.in_(nins))
            )
        ).all()
    )
    profiles = (await session.execute(select(Profile.id, Profile.email))).all()
    email_to_id = {email.lower(): pid for pid, email in profiles}

    for mapped in valid:
        assert mapped.payload is not None and mapped.nin is not None
        payload = mapped.payload
        was_existing = mapped.nin in existing_nins
        linked_id = email_to_id.get(payload.email.lower()) if payload.email else None
        fields = asdict(payload)
        if payload.joined_at is None:
            fields.pop("joined_at")
        try:
            if was_existing:
                scout = await session.scalar(
                    select(Scout).where(Scout.numero_associado == mapped.nin)
                )
                if scout is not None:
                    for attr, value in fields.items():
                        setattr(scout, attr, value)
            else:
                if linked_id:
                    fields["profile_id"] = linked_id
                session.add(Scout(**fields))
            await session.commit()
            if was_existing:
                summary["updated"] += 1
            else:
                summary["created"] += 1
                if linked_id:
                    summary["linkedToProfile"] += 1
        except IntegrityError:
            await session.rollback()
            summary["errors"].append(
                {
                    "row": mapped.row,
                    "nome": f"{payload.first_name} {payload.last_name}".strip(),
                    "error": "Conflito de unicidade (já existe outro membro)",
                }
            )

    return {"summary": summary}


@router.get("/{scout_id}")
async def get_scout(
    scout_id: str, user: CurrentUser, session: Annotated[AsyncSession, Depends(get_session)]
) -> dict[str, ScoutOut]:
    scout = await session.scalar(select(Scout).where(Scout.id == scout_id))
    if scout is None:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    return {"scout": ScoutOut.model_validate(scout)}


@router.patch("/{scout_id}")
async def update_scout(
    scout_id: str,
    admin: AdminUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    body: Annotated[dict[str, Any], Body()],
) -> dict[str, ScoutOut]:
    scout = await session.scalar(select(Scout).where(Scout.id == scout_id))
    if scout is None:
        raise HTTPException(status_code=404, detail="Membro não encontrado")

    if "firstName" in body and isinstance(body["firstName"], str):
        if not body["firstName"].strip():
            raise HTTPException(status_code=400, detail="Nome obrigatório")
        scout.first_name = body["firstName"].strip()
    if "lastName" in body and isinstance(body["lastName"], str):
        if not body["lastName"].strip():
            raise HTTPException(status_code=400, detail="Apelido obrigatório")
        scout.last_name = body["lastName"].strip()
    if "dateOfBirth" in body:
        parsed = _parse_date(body["dateOfBirth"])
        if not parsed:
            raise HTTPException(status_code=400, detail="Data de nascimento inválida")
        scout.date_of_birth = parsed
    if "section" in body:
        raw = body["section"]
        if raw is None or raw == "":
            scout.section = None
        elif raw in _VALID_SECTIONS:
            scout.section = OrdemSection(raw)
        else:
            raise HTTPException(status_code=400, detail="Secção inválida")
    if "joinedAt" in body:
        parsed = _parse_date(body["joinedAt"])
        if not parsed:
            raise HTTPException(status_code=400, detail="Data de admissão inválida")
        scout.joined_at = parsed
    if "active" in body:
        scout.active = bool(body["active"])
    if "noitesCampoInicial" in body:
        try:
            n = float(body["noitesCampoInicial"])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Noites de campo inválido") from None
        if not math.isfinite(n) or n < 0:
            raise HTTPException(status_code=400, detail="Noites de campo inválido")
        scout.noites_campo_inicial = math.floor(n)
    for camel, attr in _OPTIONAL_FIELDS.items():
        if camel in body:
            setattr(scout, attr, _optional_string(body[camel]))

    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=409, detail="Nº de associado já existe") from None
    await session.refresh(scout)
    return {"scout": ScoutOut.model_validate(scout)}


@router.delete("/{scout_id}")
async def delete_scout(
    scout_id: str, admin: AdminUser, session: Annotated[AsyncSession, Depends(get_session)]
) -> dict[str, bool]:
    scout = await session.scalar(select(Scout).where(Scout.id == scout_id))
    if scout is None:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    await session.delete(scout)
    await session.commit()
    return {"ok": True}


@router.get("/{scout_id}/nights-badges")
async def get_nights_badges(
    scout_id: str, user: CurrentUser, session: Annotated[AsyncSession, Depends(get_session)]
) -> dict[str, list[NightsBadgeOut]]:
    rows = (
        await session.scalars(
            select(ScoutNightsBadge)
            .where(ScoutNightsBadge.scout_id == scout_id)
            .order_by(ScoutNightsBadge.count.asc())
        )
    ).all()
    return {"badges": [NightsBadgeOut.model_validate(b) for b in rows]}


@router.put("/{scout_id}/nights-badges")
async def put_nights_badges(
    scout_id: str,
    admin: AdminUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    body: Annotated[dict[str, Any], Body()],
) -> dict[str, Any]:
    scout = await session.scalar(select(Scout.id).where(Scout.id == scout_id))
    if scout is None:
        raise HTTPException(status_code=404, detail="Membro não encontrado")

    badges = body.get("badges")
    if not isinstance(badges, list):
        raise HTTPException(status_code=400, detail="Formato inválido")

    to_delete: list[int] = []
    to_upsert: list[tuple[int, datetime]] = []
    for raw in badges:
        if raw is None or not isinstance(raw, dict):
            continue
        count: Any = raw.get("count")
        if not is_nights_badge_count(count):
            raise HTTPException(status_code=400, detail=f"count inválido ({count})")
        awarded = raw.get("awardedAt")
        if awarded is None or awarded == "":
            to_delete.append(int(count))
            continue
        parsed = _parse_date(awarded)
        if not parsed:
            raise HTTPException(status_code=400, detail=f"Data inválida para {count}")
        to_upsert.append((int(count), parsed))

    if to_delete:
        await session.execute(
            delete(ScoutNightsBadge).where(
                ScoutNightsBadge.scout_id == scout_id,
                ScoutNightsBadge.count.in_(to_delete),
            )
        )
    for count, awarded_at in to_upsert:
        stmt = (
            pg_insert(ScoutNightsBadge)
            .values(scout_id=scout_id, count=count, awarded_at=awarded_at)
            .on_conflict_do_update(
                index_elements=[ScoutNightsBadge.scout_id, ScoutNightsBadge.count],
                set_={"awarded_at": awarded_at},
            )
        )
        await session.execute(stmt)
    await session.commit()

    rows = (
        await session.scalars(
            select(ScoutNightsBadge)
            .where(ScoutNightsBadge.scout_id == scout_id)
            .order_by(ScoutNightsBadge.count.asc())
        )
    ).all()
    return {
        "badges": [NightsBadgeOut.model_validate(b) for b in rows],
        "allCounts": list(NIGHTS_BADGE_COUNTS),
    }
