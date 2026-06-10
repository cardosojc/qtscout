import io
from datetime import UTC, datetime
from typing import Annotated, Any

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.ordem_categories import (
    ItemShape,
    is_ordem_section,
    validate_item_data,
)
from app.core.ordem_permissions import ProfileForAuth, can_manage_item, resolve_category
from app.core.ordem_resolver import build_display, resolve_refs
from app.core.siie_atividades_import import map_activity_row
from app.db import get_session
from app.deps import AdminUser, CurrentUser
from app.models import OrdemItem, Profile, Scout
from app.models.enums import OrdemSection
from app.schemas.ordem_item import OrdemItemOut

router = APIRouter(prefix="/ordem-items", tags=["ordem-items"])


def _parse_dt(value: Any) -> datetime | None:
    if not isinstance(value, str) or not value:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.astimezone(UTC).replace(tzinfo=None) if parsed.tzinfo else parsed


async def _load_auth(session: AsyncSession, user_id: str) -> ProfileForAuth | None:
    row = (
        await session.execute(
            select(Profile.role, Profile.roles, Profile.section).where(Profile.id == user_id)
        )
    ).first()
    if row is None:
        return None
    return ProfileForAuth(
        role=str(row.role), roles=list(row.roles), section=str(row.section) if row.section else None
    )


async def _validate_refs(
    session: AsyncSession, shape: ItemShape, value: dict[str, Any], section: str | None
) -> str | None:
    scout_ids: list[str] = []
    profile_ids: list[str] = []
    if shape == "MEMBER_REF" and isinstance(value.get("scoutId"), str):
        scout_ids.append(value["scoutId"])
    if shape == "NOITES_REF" and isinstance(value.get("scoutIds"), list):
        scout_ids.extend(i for i in value["scoutIds"] if isinstance(i, str))
    if shape == "PROFILE_REF" and isinstance(value.get("profileId"), str):
        profile_ids.append(value["profileId"])
    if shape == "SCOUT_OR_PROFILE_REF" and isinstance(value.get("refId"), str):
        if value.get("kind") == "scout":
            scout_ids.append(value["refId"])
        if value.get("kind") == "profile":
            profile_ids.append(value["refId"])

    if scout_ids:
        found = (
            await session.execute(
                select(Scout.id, Scout.section).where(Scout.id.in_(scout_ids))
            )
        ).all()
        if len(found) != len(scout_ids):
            return "Membro não encontrado"
        if section and any(s.section != section for s in found):
            return "O membro selecionado pertence a outra secção"
    if profile_ids:
        count = await session.scalar(
            select(func.count()).select_from(Profile).where(Profile.id.in_(profile_ids))
        )
        if count != len(profile_ids):
            return "Dirigente não encontrado"
    return None


def _annotated(item: OrdemItem, refs: Any) -> OrdemItemOut:
    out = OrdemItemOut.model_validate(item)
    out.data = {**(item.data or {}), "_display": build_display(item.data, refs)}
    return out


@router.get("")
async def list_items(
    user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    from_: Annotated[str | None, Query(alias="from")] = None,
    to: Annotated[str | None, Query()] = None,
    section: Annotated[str | None, Query()] = None,
    category: Annotated[str | None, Query()] = None,
    included: Annotated[str | None, Query()] = None,
) -> dict[str, list[OrdemItemOut]]:
    conditions = []
    if from_:
        conditions.append(OrdemItem.date >= _require_dt(from_))
    if to:
        conditions.append(OrdemItem.date <= _require_dt(to))
    if section and is_ordem_section(section):
        conditions.append(OrdemItem.section == OrdemSection(section))
    if category:
        conditions.append(OrdemItem.category == category)
    if included == "true":
        conditions.append(OrdemItem.included_in_os_id.is_not(None))
    if included == "false":
        conditions.append(OrdemItem.included_in_os_id.is_(None))

    items = (
        await session.scalars(
            select(OrdemItem)
            .where(*conditions)
            .options(selectinload(OrdemItem.created_by))
            .order_by(OrdemItem.date.desc(), OrdemItem.created_at.desc())
        )
    ).all()
    refs = await resolve_refs(session, items)
    return {"items": [_annotated(item, refs) for item in items]}


def _require_dt(value: str) -> datetime:
    parsed = _parse_dt(value)
    if parsed is None:
        raise HTTPException(status_code=400, detail="Data inválida")
    return parsed


@router.post("")
async def create_item(
    user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    body: Annotated[dict[str, Any], Body()],
) -> dict[str, OrdemItemOut]:
    raw_category = body.get("category")
    category = resolve_category(raw_category) if isinstance(raw_category, str) else None
    if category is None:
        raise HTTPException(status_code=400, detail="Categoria inválida")

    raw_section = body.get("section")
    if raw_section is None:
        section: str | None = None
    elif is_ordem_section(raw_section):
        section = raw_section
    else:
        raise HTTPException(status_code=400, detail="Secção inválida")
    if category.scope == "SECTION" and not section:
        raise HTTPException(status_code=400, detail="Secção é obrigatória")
    if category.scope == "GROUP" and section:
        raise HTTPException(status_code=400, detail="Categoria de grupo não aceita secção")

    date = _parse_dt(body.get("date"))
    if date is None:
        raise HTTPException(status_code=400, detail="Data inválida")

    result = validate_item_data(category.shape, body.get("data"))
    if not result.ok or result.value is None:
        raise HTTPException(status_code=400, detail=result.error or "Dados inválidos")

    profile = await _load_auth(session, user.id)
    if profile is None:
        raise HTTPException(status_code=404, detail="Perfil não encontrado")
    if not can_manage_item(profile, category, section):
        raise HTTPException(status_code=403, detail="Sem permissões para esta categoria/secção")

    ref_error = await _validate_refs(session, category.shape, result.value, section)
    if ref_error:
        raise HTTPException(status_code=400, detail=ref_error)

    item = OrdemItem(
        category=category.key,
        section=OrdemSection(section) if section else None,
        date=date,
        data=result.value,
        created_by_id=user.id,
    )
    session.add(item)
    await session.commit()
    created = await session.scalar(
        select(OrdemItem).where(OrdemItem.id == item.id).options(selectinload(OrdemItem.created_by))
    )
    assert created is not None
    return {"item": OrdemItemOut.model_validate(created)}


@router.post("/import-activities")
async def import_activities(
    admin: AdminUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    file: Annotated[UploadFile | None, File()] = None,
) -> dict[str, Any]:
    if file is None:
        raise HTTPException(status_code=400, detail="Ficheiro não fornecido")
    try:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        if not workbook.worksheets:
            raise HTTPException(status_code=400, detail="Folha vazia")
        sheet = workbook.worksheets[0]
        it = sheet.iter_rows(values_only=True)
        header = list(next(it, []))
        records = [
            dict(zip(header, row, strict=False)) for row in it if any(v is not None for v in row)
        ]
    except HTTPException:
        raise
    except Exception as err:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Não foi possível ler o ficheiro: {err}"
        ) from err

    summary: dict[str, Any] = {
        "total": len(records),
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "errors": [],
    }

    valid = []
    for i, record in enumerate(records):
        mapped = map_activity_row(record, i + 2)
        if not mapped.ok:
            summary["errors"].append(
                {"row": mapped.row, "descricao": mapped.descricao, "error": mapped.error}
            )
        else:
            valid.append(mapped)

    if not valid:
        return {"summary": summary}

    ext_ids = [m.payload.external_id for m in valid if m.payload]
    existing_rows = (
        await session.execute(
            select(OrdemItem.external_id, OrdemItem.included_in_os_id).where(
                OrdemItem.external_id.in_(ext_ids)
            )
        )
    ).all()
    included_map = {r.external_id: r.included_in_os_id for r in existing_rows}

    for mapped in valid:
        payload = mapped.payload
        assert payload is not None
        was_existing = payload.external_id in included_map
        if was_existing and included_map[payload.external_id]:
            summary["skipped"] += 1
            continue
        data = {"nome": payload.nome, "datas": payload.datas, "local": payload.local}
        try:
            if was_existing:
                item = await session.scalar(
                    select(OrdemItem).where(OrdemItem.external_id == payload.external_id)
                )
                if item is not None:
                    item.category = "ATIVIDADE"
                    item.section = OrdemSection(payload.section) if payload.section else None
                    item.date = payload.date
                    item.data = data
                summary["updated"] += 1
            else:
                session.add(
                    OrdemItem(
                        external_id=payload.external_id,
                        category="ATIVIDADE",
                        section=OrdemSection(payload.section) if payload.section else None,
                        date=payload.date,
                        data=data,
                        created_by_id=admin.id,
                    )
                )
                summary["created"] += 1
            await session.commit()
        except IntegrityError:
            await session.rollback()
            if was_existing:
                summary["updated"] -= 1
            else:
                summary["created"] -= 1
            summary["errors"].append(
                {"row": mapped.row, "descricao": payload.nome, "error": "Conflito de unicidade"}
            )

    return {"summary": summary}


@router.patch("/{item_id}")
async def update_item(
    item_id: str,
    user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    body: Annotated[dict[str, Any], Body()],
) -> dict[str, OrdemItemOut]:
    existing = await session.scalar(select(OrdemItem).where(OrdemItem.id == item_id))
    if existing is None:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    if existing.included_in_os_id:
        raise HTTPException(status_code=409, detail="Item já incluído numa Ordem de Serviço")

    category = resolve_category(existing.category)
    if category is None:
        raise HTTPException(status_code=500, detail="Categoria inválida")

    profile = await _load_auth(session, user.id)
    if profile is None:
        raise HTTPException(status_code=404, detail="Perfil não encontrado")
    if existing.created_by_id != user.id and profile.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Sem permissões")

    if "date" in body:
        date = _parse_dt(body["date"])
        if date is None:
            raise HTTPException(status_code=400, detail="Data inválida")
        existing.date = date
    if "data" in body:
        result = validate_item_data(category.shape, body["data"])
        if not result.ok or result.value is None:
            raise HTTPException(status_code=400, detail=result.error or "Dados inválidos")
        existing.data = result.value
    if "section" in body:
        if category.scope == "GROUP":
            raise HTTPException(status_code=400, detail="Categoria de grupo não aceita secção")
        if not is_ordem_section(body["section"]):
            raise HTTPException(status_code=400, detail="Secção inválida")
        if not can_manage_item(profile, category, body["section"]):
            raise HTTPException(status_code=403, detail="Sem permissões para essa secção")
        existing.section = OrdemSection(body["section"])

    await session.commit()
    updated = await session.scalar(
        select(OrdemItem)
        .where(OrdemItem.id == item_id)
        .options(selectinload(OrdemItem.created_by))
    )
    assert updated is not None
    return {"item": OrdemItemOut.model_validate(updated)}


@router.delete("/{item_id}")
async def delete_item(
    item_id: str, user: CurrentUser, session: Annotated[AsyncSession, Depends(get_session)]
) -> dict[str, bool]:
    existing = await session.scalar(select(OrdemItem).where(OrdemItem.id == item_id))
    if existing is None:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    if existing.included_in_os_id:
        raise HTTPException(status_code=409, detail="Item já incluído numa Ordem de Serviço")
    profile = await _load_auth(session, user.id)
    if profile is None:
        raise HTTPException(status_code=404, detail="Perfil não encontrado")
    if existing.created_by_id != user.id and profile.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Sem permissões")
    await session.delete(existing)
    await session.commit()
    return {"ok": True}
